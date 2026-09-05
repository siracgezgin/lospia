#!/usr/bin/env python3
"""Excel çalışma kitabını AF Teamwork'e TABLO olarak aktarır (görselleriyle).

Sıraç (2026-09-06): "Excelleri de oluşturalım /documents kısmında. Yani ordaki
format birebir burada olmalı artık." Ve daha önce: "Aslı Hanım nasıl yapmıştı,
aşina olduğu düzende olsun ki yabancılık çekmesin."

BU YÜZDEN: bir Excel DOSYASI → bir tablo, her Excel SAYFASI → o tablonun bir
SEKMESİ. Sekme adları, sütun sırası ve sütun genişlikleri kaynaktaki gibi
kalır. Yeni bir düzen icat edilmez.

GÖRSELLER YENİDEN YÜKLENMEZ. import-excel-images.py onları Drive'a çoktan
koydu; burada aynı SHA-256 özeti hesaplanıp Drive'daki kaydın kimliği bulunur
ve hücreye REFERANS yazılır (lib/sheets/model → CellImage). Yani aynı fotoğraf
Drive'da tek kopya durur, tabloda görünür. Önce görsel betiğini çalıştırın.

KULLANIM
  python3 scripts/import-excel-sheets.py --file "sonh/AFR-AF  (2).xlsx" --dry-run
  python3 scripts/import-excel-sheets.py --file "sonh/AFR-AF  (2).xlsx"

  Canlı:
    IMPORT_SUPABASE_URL=https://<proj>.supabase.co \
    IMPORT_SUPABASE_SERVICE_ROLE_KEY=sb_secret_… \
    python3 scripts/import-excel-sheets.py --file "sonh/AFR-AF  (2).xlsx" --prod

İDEMPOTENT: tablo BAŞLIĞA göre bulunur; varsa anlık görüntüsü güncellenir,
yoksa oluşturulur. Tekrar çalıştırmak kopya üretmez.
"""
import argparse
import hashlib
import json
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
import datetime

try:
    import openpyxl
except ImportError:
    sys.exit("❌  openpyxl gerekli:  python3 -m pip install openpyxl")

# lib/sheets/model.ts ile AYNI sınırlar — aşan sayfa kırpılır ve söylenir.
MAX_ROWS = 5000
MAX_COLS = 100
DEFAULT_COL_W = 128
# Excel sütun genişliği "karakter" birimindedir; ~7 piksel/karakter yaklaşımı
# Excel'in kendi dönüşümüne yakın sonuç veriyor.
PX_PER_CHAR = 7


class Api:
    def __init__(self, url, key):
        self.url = url.rstrip("/")
        self.key = key

    def _req(self, method, path, body=None, headers=None):
        h = {"apikey": self.key, "Authorization": f"Bearer {self.key}"}
        if headers:
            h.update(headers)
        data = None
        if body is not None:
            data = json.dumps(body).encode()
            h.setdefault("Content-Type", "application/json")
        req = urllib.request.Request(self.url + path, data=data, headers=h, method=method)
        try:
            with urllib.request.urlopen(req) as resp:
                payload = resp.read()
                return json.loads(payload) if payload else None
        except urllib.error.HTTPError as e:
            raise RuntimeError(f"{method} {path} → {e.code}: {e.read().decode('utf-8', 'replace')[:400]}") from None

    def get(self, p):
        return self._req("GET", p)

    def post(self, p, b, h=None):
        return self._req("POST", p, b, h)

    def patch(self, p, b, h=None):
        return self._req("PATCH", p, b, h)


def load_env_local():
    env = {}
    p = os.path.join(os.getcwd(), ".env.local")
    if not os.path.exists(p):
        return env
    for line in open(p, encoding="utf8"):
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k.strip()] = v.strip().strip("\"'")
    return env


PROD_ENV_FILE = ".env.prod.local"


def load_env_file(filename):
    """Basit .env okuyucu — yalnız KEY=VALUE satırları."""
    env = {}
    p = os.path.join(os.getcwd(), filename)
    if not os.path.exists(p):
        return env
    for line in open(p, encoding="utf8"):
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k.strip()] = v.strip().strip("\"'")
    return env


def prod_credentials():
    """Canlı kimlik bilgileri: önce ortam değişkeni, sonra .env.prod.local.

    Sıraç (2026-09-06) üç kez üst üste kabuk değişkenlerine takıldı: komut
    satırındaki yer tutucular aynen yapıştırıldı, `export` boş kaldı. Anahtarı
    bir kez dosyaya yazmak bu sınıfı tamamen kapatıyor. Dosya .gitignore'da
    (.env.* kalıbı) — depoya giremez.
    """
    url = os.environ.get("IMPORT_SUPABASE_URL")
    key = os.environ.get("IMPORT_SUPABASE_SERVICE_ROLE_KEY")
    if url and key:
        return url, key
    f = load_env_file(PROD_ENV_FILE)
    return url or f.get("IMPORT_SUPABASE_URL"), key or f.get("IMPORT_SUPABASE_SERVICE_ROLE_KEY")


def resolve_target(prod):
    url, key = prod_credentials()
    if prod or url or key:
        if not url or not key:
            sys.exit(
                "\u274c  Canl\u0131 kimlik bilgileri bulunamad\u0131.\n"
                "\n    EN KOLAY YOL \u2014 proje k\u00f6k\u00fcnde " + PROD_ENV_FILE + " dosyas\u0131 olu\u015fturun:\n"
                "        IMPORT_SUPABASE_URL=https://<proje>.supabase.co\n"
                "        IMPORT_SUPABASE_SERVICE_ROLE_KEY=sb_secret_...\n"
                "\n    De\u011ferler: Panel \u2192 Settings \u2192 API \u2192 Project URL ve Secret keys \u2192 default\n"
                "    (g\u00f6z simgesine bas\u0131p a\u00e7\u0131n, sonra kopyalay\u0131n).\n"
                "\n    Bu dosya .gitignore'dad\u0131r, depoya giremez. Kabuk de\u011fi\u015fkeni de\n"
                "    \u00e7al\u0131\u015f\u0131r ama her seferinde yeniden yazmak gerekir."
            )
        if not url.startswith(("http://", "https://")):
            sys.exit(f'❌  IMPORT_SUPABASE_URL geçerli bir adres değil: "{url}"')
        if key.startswith("sb_publishable_"):
            sys.exit("❌  Bu bir tarayıcı anahtarı (publishable) — yazamaz.")
        return url, key, "PRODUCTION"
    env = load_env_local()
    url, key = env.get("NEXT_PUBLIC_SUPABASE_URL"), env.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        sys.exit("❌  .env.local içinde NEXT_PUBLIC_SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY yok")
    return url, key, "LOCAL"


def cell_text(v):
    """Hücre değeri → tabloda saklanacak metin.

    Sayılar Excel'in gösterdiği gibi değil HAM hâliyle yazılır; tablo motoru
    biçimlendirmeyi kendi yapar. Tarihler ISO'ya çevrilir — "2026-06-10
    00:00:00" gibi bir metin tabloda tarih olarak okunamazdı.
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return "DOĞRU" if v else "YANLIŞ"
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip() or None


def main():
    ap = argparse.ArgumentParser(description="Excel kitabını AF Teamwork tablosuna aktarır")
    ap.add_argument("--file", required=True)
    ap.add_argument("--prod", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--title", default=None, help="Tablo başlığı (varsayılan: dosya adı)")
    ap.add_argument("--folder", default=None, help="Hedef klasör adı (varsayılan: kök)")
    args = ap.parse_args()

    if not os.path.exists(args.file):
        sys.exit(f"❌  Dosya bulunamadı: {args.file}")

    title = args.title or re.sub(r"\.xlsx?$", "", os.path.basename(args.file)).strip()
    print(f"📄  Kaynak: {os.path.basename(args.file)}")
    print(f"📊  Tablo başlığı: {title}")
    print("    Okunuyor… (gömülü görselli büyük dosyalarda birkaç dakika sürebilir)")

    wb = openpyxl.load_workbook(args.file, data_only=True)

    # ── Görsel çapaları: (sayfa, satır, sütun) → SHA-256 özeti ───────────────
    # Görseller ZATEN Drive'da; burada yalnız hangi hücreye denk geldiklerini
    # ve hangi Drive kaydına karşılık geldiklerini buluyoruz.
    anchors = {}
    for name in wb.sheetnames:
        for im in getattr(wb[name], "_images", []):
            try:
                data = im._data()
            except Exception:
                continue
            if not data:
                continue
            fmt = (getattr(im, "format", None) or "png").lower()
            ext = "jpg" if fmt in ("jpg", "jpeg") else fmt
            try:
                r, c = im.anchor._from.row, im.anchor._from.col
            except Exception:
                continue
            anchors.setdefault(name, {})[(r, c)] = f"{hashlib.sha256(data).hexdigest()[:16]}.{ext}"

    sheets = []
    truncated = []
    image_slots = 0
    for name in wb.sheetnames:
        ws = wb[name]
        max_r = min(ws.max_row or 1, MAX_ROWS)
        max_c = min(ws.max_column or 1, MAX_COLS)
        if (ws.max_row or 0) > MAX_ROWS or (ws.max_column or 0) > MAX_COLS:
            truncated.append(f"{name} ({ws.max_row}×{ws.max_column} → {max_r}×{max_c})")

        cells = {}
        for row in ws.iter_rows(min_row=1, max_row=max_r, max_col=max_c):
            for cell in row:
                text = cell_text(cell.value)
                if text is None:
                    continue
                cells[f"{cell.row - 1}:{cell.column - 1}"] = {"v": text}

        # Sütun genişlikleri — Aslı Hanım'ın düzeni bozulmasın.
        col_w = {}
        for letter, dim in (ws.column_dimensions or {}).items():
            if not dim.width:
                continue
            try:
                idx = openpyxl.utils.column_index_from_string(letter) - 1
            except Exception:
                continue
            if 0 <= idx < MAX_COLS:
                col_w[str(idx)] = max(48, min(400, round(dim.width * PX_PER_CHAR)))

        # Birleştirilmiş hücreler — "r1:c1:r2:c2"
        merges = []
        for rng in (ws.merged_cells.ranges if ws.merged_cells else []):
            if rng.min_row - 1 >= max_r or rng.min_col - 1 >= max_c:
                continue
            merges.append(f"{rng.min_row - 1}:{rng.min_col - 1}:"
                          f"{min(rng.max_row, max_r) - 1}:{min(rng.max_col, max_c) - 1}")

        sheets.append({
            "name": name[:60], "rows": max(max_r, 30), "cols": max(max_c, 12),
            "cells": cells, "colW": col_w, "merges": merges,
            "_anchors": anchors.get(name, {}),
        })
        image_slots += len(anchors.get(name, {}))

    print(f"🔎  {len(sheets)} sayfa · {sum(len(s['cells']) for s in sheets)} dolu hücre "
          f"· {image_slots} görsel yerleşimi")
    if truncated:
        print(f"    ⚠  Sınır aşıldı, kırpılan sayfalar: {', '.join(truncated)}")

    if args.dry_run:
        print("\n— ÖNİZLEME (hiçbir şey yazılmadı) —")
        for s in sheets[:40]:
            print(f"    {s['name']}: {s['rows']}×{s['cols']} · {len(s['cells'])} hücre "
                  f"· {len(s['_anchors'])} görsel")
        return

    url, key, label = resolve_target(args.prod)
    print(f"🎯  Hedef: {label}")
    api = Api(url, key)

    ws_rows = api.get("/rest/v1/workspaces?select=id,name")
    if not ws_rows:
        sys.exit("❌  Çalışma alanı bulunamadı. Anahtar RLS'i aşmıyor olabilir.")
    wsr = next((w for w in ws_rows if w["name"] == "AF Operasyon"), ws_rows[0])
    workspace_id = wsr["id"]
    print(f"✅  Çalışma alanı: {wsr['name']} ({workspace_id})")

    owner = api.get(f"/rest/v1/workspace_members?select=user_id&workspace_id=eq.{workspace_id}"
                    f"&role=eq.owner&limit=1")
    created_by = owner[0]["user_id"] if owner else None

    # ── Görsel adı → Drive kaydı ─────────────────────────────────────────────
    # Tek turda çekilir; görsel başına sorgu atmak 1000 istek olurdu.
    drive = {}
    offset = 0
    while True:
        page = api.get(f"/rest/v1/operation_documents?select=id,file_path,file_name"
                       f"&workspace_id=eq.{workspace_id}&document_type=eq.file"
                       f"&limit=1000&offset={offset}")
        if not page:
            break
        for row in page:
            fp = row.get("file_path") or ""
            if fp:
                drive[fp.rsplit("/", 1)[-1]] = {"id": row["id"], "name": row.get("file_name")}
        if len(page) < 1000:
            break
        offset += 1000
    print(f"🖼   Drive'da {len(drive)} görsel kaydı bulundu.")

    folder_id = None
    if args.folder:
        found = api.get(f"/rest/v1/document_folders?select=id&workspace_id=eq.{workspace_id}"
                        f"&name=eq.{urllib.parse.quote(args.folder)}&limit=1")
        if found:
            folder_id = found[0]["id"]
        else:
            made = api.post("/rest/v1/document_folders", {
                "workspace_id": workspace_id, "name": args.folder, "parent_id": None,
                "visibility": "all", "created_by": created_by,
            }, {"Prefer": "return=representation"})
            folder_id = made[0]["id"]
        print(f"📁  Klasör: {args.folder}")

    # Çapaları hücre referansına çevir.
    linked = missing = 0
    for i, s in enumerate(sheets):
        for (r, c), fname in s.pop("_anchors").items():
            if r >= s["rows"] or c >= s["cols"]:
                continue
            hit = drive.get(fname)
            if not hit:
                missing += 1
                continue
            k = f"{r}:{c}"
            cell = s["cells"].get(k, {})
            cell["img"] = {"id": hit["id"], "name": hit["name"] or fname}
            s["cells"][k] = cell
            linked += 1
        s["id"] = f"s{i + 1}"

    print(f"🔗  {linked} hücreye görsel bağlandı" + (f" · {missing} görsel Drive'da bulunamadı" if missing else ""))
    if missing:
        print("    (Önce scripts/import-excel-images.py çalıştırıldı mı?)")

    snapshot = {"engine": "wb", "sheets": sheets, "active": 0}

    existing = api.get(f"/rest/v1/operation_spreadsheets?select=id&workspace_id=eq.{workspace_id}"
                       f"&title=eq.{urllib.parse.quote(title)}&limit=1")
    if existing:
        api.patch(f"/rest/v1/operation_spreadsheets?id=eq.{existing[0]['id']}",
                  {"snapshot": snapshot, "folder_id": folder_id}, {"Prefer": "return=minimal"})
        print(f"\n✅  Güncellendi — {title}")
    else:
        api.post("/rest/v1/operation_spreadsheets", {
            "workspace_id": workspace_id,
            "title": title,
            "sheet_type": "freeform",
            "snapshot": snapshot,
            "status": "active",
            "section": "teamwork",
            "visibility": "all",
            "folder_id": folder_id,
            "owner_id": created_by,
            "created_by": created_by,
        }, {"Prefer": "return=minimal"})
        print(f"\n✅  Oluşturuldu — {title}")
    print(f"    {len(sheets)} sekme · {sum(len(s['cells']) for s in sheets)} hücre · {linked} görsel")


if __name__ == "__main__":
    main()
