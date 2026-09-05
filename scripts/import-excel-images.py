#!/usr/bin/env python3
"""Excel'e gömülü görselleri Drive'a (AF Teamwork) klasörleyerek aktarır.

Sıraç (2026-09-06): "Aynı resmi birkaç defa yüklemek sistemi gereksiz
ağırlaştırır… önce klasör oluşturup oraya ekleyelim, sonra excelde o resmi
klasörden çağırtıp gösterelim."

NE YAPAR
  1. Çalışma kitabındaki her sayfanın gömülü görsellerini çıkarır.
  2. Sayfa ADINA göre bir klasör açar (Aslı Hanım bugün sekme adıyla geziniyor;
     Drive'da da aynı adı görsün).
  3. Görseli `documents` kovasına yükler ve `operation_documents`'a bir satır
     yazar — yani Drive'da normal bir dosya olur.
  4. Yanına KÜÇÜK BİR ÖNİZLEME üretir (`thumbs/…`, thumb_path). Tablo hücresinde
     bu gösterilir: 498 KB yerine ~15 KB iner.

ÖNİZLEME NEDEN ŞART: ücretsiz planın aylık 5 GB indirme hakkı var. Ortalama
498 KB'lık görsellerle dolu bir tabloyu bir kez açmak 25 MB indiriyor; aynı
sayfa önizlemelerle 0,75 MB ediyor.

ORİJİNAL KORUNUR: varsayılan olarak görsel OLDUĞU GİBİ yüklenir — Excel'deki
kopya tek nüsha, onu küçültmek geri alınamaz. Yer sıkıntısı olursa
`--shrink 1600` ile uzun kenar sınırlanabilir; bu BİLİNÇLİ bir karardır.

KÜÇÜLTME macOS'un kendi `sips` aracıyla yapılır — yeni bağımlılık yok.

KULLANIM
  python3 scripts/import-excel-images.py --file "sonh/AFR-AF  (2).xlsx" --dry-run
  python3 scripts/import-excel-images.py --file "sonh/AFR-AF  (2).xlsx"

  Canlı:
    IMPORT_SUPABASE_URL=https://<proj>.supabase.co \
    IMPORT_SUPABASE_SERVICE_ROLE_KEY=sb_secret_… \
    python3 scripts/import-excel-images.py --file "sonh/AFR-AF  (2).xlsx" --prod

İDEMPOTENT: her görsel içeriğinin SHA-256'sıyla anahtarlanır (dosya adına
yazılır). Aynı görsel ikinci kez yüklenmez; betik tekrar çalıştırılabilir.
"""
import argparse
import hashlib
import json
import mimetypes
import os
import re
import subprocess
import sys
import tempfile
import urllib.error
import urllib.parse
import urllib.request

try:
    import openpyxl
except ImportError:
    sys.exit("❌  openpyxl gerekli:  python3 -m pip install openpyxl")

BUCKET = "documents"
# Önizleme, çalışma alanı klasörünün İÇİNDE durur. Depolama politikası
# (20240312) ilk yol parçasının çalışma alanı kimliği olmasını şart koşuyor:
#   is_workspace_member((storage.foldername(name))[1]::uuid)
# İlk sürümde önizleme "thumbs/<ws>/…" idi; ilk parça "thumbs" olduğu için
# politika geçmiyor, imza ÜRETİLEMİYOR ve seçicide bütün görseller kırık
# görünüyordu (Sıraç, 2026-09-06). Drive listesi çalışıyordu çünkü o orijinali
# imzalıyor. Doğru yol: <ws>/thumbs/<özet>.<uzantı>
THUMB_DIR = "thumbs"
LEGACY_THUMB_PREFIX = "thumbs"
THUMB_MAX_PX = 400
ROOT_FOLDER_DEFAULT = "Excel Görselleri"

# Windows'ta yasak adlar ve karakterler — arşiv/indirme sırasında bozulmasın.
BAD_CHARS = re.compile(r'[\\/:*?"<>|\x00-\x1f]')
RESERVED = {"CON", "PRN", "AUX", "NUL", *(f"COM{i}" for i in range(1, 10)),
            *(f"LPT{i}" for i in range(1, 10))}


def safe_name(name: str, limit: int = 90) -> str:
    """Klasör/dosya adı olarak güvenli hâle getirir; Türkçe harfleri korur."""
    out = BAD_CHARS.sub(" ", str(name or "")).strip().strip(".")
    out = re.sub(r"\s+", " ", out)
    if out.upper() in RESERVED:
        out = f"_{out}"
    return (out[:limit].strip() or "Adsız")


# ── HTTP ─────────────────────────────────────────────────────────────────────

class Api:
    def __init__(self, url: str, key: str):
        self.url = url.rstrip("/")
        self.key = key

    def _req(self, method, path, body=None, headers=None, raw=False):
        h = {"apikey": self.key, "Authorization": f"Bearer {self.key}"}
        if headers:
            h.update(headers)
        data = body
        if body is not None and not raw:
            data = json.dumps(body).encode()
            h.setdefault("Content-Type", "application/json")
        req = urllib.request.Request(self.url + path, data=data, headers=h, method=method)
        try:
            with urllib.request.urlopen(req) as resp:
                payload = resp.read()
                if not payload:
                    return None
                try:
                    return json.loads(payload)
                except json.JSONDecodeError:
                    return payload
        except urllib.error.HTTPError as e:
            detail = e.read().decode("utf-8", "replace")[:400]
            raise RuntimeError(f"{method} {path} → {e.code}: {detail}") from None

    def get(self, path):
        return self._req("GET", path)

    def post(self, path, body, headers=None):
        return self._req("POST", path, body, headers)

    def delete_objects(self, paths):
        return self._req("DELETE", f"/storage/v1/object/{BUCKET}",
                         {"prefixes": paths})

    def list_objects(self, prefix, limit=100, offset=0):
        return self._req("POST", f"/storage/v1/object/list/{BUCKET}",
                         {"prefix": prefix, "limit": limit, "offset": offset,
                          "sortBy": {"column": "name", "order": "asc"}})

    def upload(self, path, data, content_type):
        return self._req(
            "POST", f"/storage/v1/object/{BUCKET}/{urllib.parse.quote(path)}",
            body=data, headers={"Content-Type": content_type, "x-upsert": "true"}, raw=True,
        )


# ── Ortam ────────────────────────────────────────────────────────────────────

def load_env_local():
    env = {}
    p = os.path.join(os.getcwd(), ".env.local")
    if not os.path.exists(p):
        return env
    for line in open(p, encoding="utf8"):
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k.strip()] = v.strip().strip("\"'")
    return env


PROD_ENV_FILE = ".env.prod.local"


def load_env_file(filename):
    """Basit .env okuyucu — yalnız KEY=VALUE satırları."""
    env = {}
    p = os.path.join(os.getcwd(), filename)
    if not os.path.exists(p):
        return env
    for line in open(p, encoding="utf8"):
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k.strip()] = v.strip().strip("\"'")
    return env


def prod_credentials():
    """Canlı kimlik bilgileri: önce ortam değişkeni, sonra .env.prod.local.

    Sıraç (2026-09-06) üç kez üst üste kabuk değişkenlerine takıldı: komut
    satırındaki yer tutucular aynen yapıştırıldı, `export` boş kaldı. Anahtarı
    bir kez dosyaya yazmak bu sınıfı tamamen kapatıyor. Dosya .gitignore'da
    (.env.* kalıbı) — depoya giremez.
    """
    url = os.environ.get("IMPORT_SUPABASE_URL")
    key = os.environ.get("IMPORT_SUPABASE_SERVICE_ROLE_KEY")
    if url and key:
        return url, key
    f = load_env_file(PROD_ENV_FILE)
    return url or f.get("IMPORT_SUPABASE_URL"), key or f.get("IMPORT_SUPABASE_SERVICE_ROLE_KEY")


def resolve_target(prod: bool):
    url, key = prod_credentials()
    if prod or url or key:
        if not url or not key:
            sys.exit(
                "\u274c  Canl\u0131 kimlik bilgileri bulunamad\u0131.\n"
                "\n    EN KOLAY YOL \u2014 proje k\u00f6k\u00fcnde " + PROD_ENV_FILE + " dosyas\u0131 olu\u015fturun:\n"
                "        IMPORT_SUPABASE_URL=https://<proje>.supabase.co\n"
                "        IMPORT_SUPABASE_SERVICE_ROLE_KEY=sb_secret_...\n"
                "\n    De\u011ferler: Panel \u2192 Settings \u2192 API \u2192 Project URL ve Secret keys \u2192 default\n"
                "    (g\u00f6z simgesine bas\u0131p a\u00e7\u0131n, sonra kopyalay\u0131n).\n"
                "\n    Bu dosya .gitignore'dad\u0131r, depoya giremez. Kabuk de\u011fi\u015fkeni de\n"
                "    \u00e7al\u0131\u015f\u0131r ama her seferinde yeniden yazmak gerekir."
            )
        if not url.startswith(("http://", "https://")):
            sys.exit(f'❌  IMPORT_SUPABASE_URL geçerli bir adres değil: "{url}"\n'
                     "    https://<proje>.supabase.co biçiminde olmalı.")
        if key.startswith("sb_publishable_"):
            sys.exit("❌  Bu bir tarayıcı anahtarı (publishable) — yazamaz.\n"
                     "    Secret keys → default (sb_secret_…) kullanın.")
        return url, key, "PRODUCTION"
    env = load_env_local()
    url = env.get("NEXT_PUBLIC_SUPABASE_URL")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        sys.exit("❌  .env.local içinde NEXT_PUBLIC_SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY yok")
    return url, key, "LOCAL"


# ── Görsel işleme ────────────────────────────────────────────────────────────

def resize(data: bytes, ext: str, max_px: int):
    """`sips` ile uzun kenarı max_px'e indirir. Başarısız olursa None döner ve
       çağıran orijinali kullanır — küçültememek veri kaybı sebebi olmamalı."""
    try:
        with tempfile.TemporaryDirectory() as tmp:
            src = os.path.join(tmp, f"in.{ext}")
            dst = os.path.join(tmp, f"out.{ext}")
            with open(src, "wb") as f:
                f.write(data)
            r = subprocess.run(
                ["sips", "--resampleHeightWidthMax", str(max_px), src, "--out", dst],
                capture_output=True, timeout=60,
            )
            if r.returncode != 0 or not os.path.exists(dst):
                return None
            with open(dst, "rb") as f:
                return f.read()
    except Exception:
        return None


def sheet_images(ws):
    """Sayfadaki gömülü görseller. openpyxl'de _data() bir kez okunabilir."""
    out = []
    for im in getattr(ws, "_images", []):
        try:
            data = im._data()
        except Exception:
            continue
        if not data:
            continue
        fmt = (getattr(im, "format", None) or "png").lower()
        if fmt == "jpg":
            fmt = "jpeg"
        try:
            row = im.anchor._from.row + 1
            col = im.anchor._from.col + 1
        except Exception:
            row, col = 0, 0
        out.append({"data": data, "ext": fmt, "row": row, "col": col})
    return out


# ── Ana akış ─────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="Excel görsellerini Drive'a aktarır")
    ap.add_argument("--file", required=True, help="Kaynak .xlsx")
    ap.add_argument("--prod", action="store_true")
    ap.add_argument("--dry-run", action="store_true", help="Hiçbir şey yazmaz")
    ap.add_argument("--root", default=ROOT_FOLDER_DEFAULT, help="Kök klasör adı")
    ap.add_argument("--drop-legacy-thumbs", action="store_true",
                    help="Eski 'thumbs/<ws>/…' yolundaki sahipsiz önizlemeleri siler")
    ap.add_argument("--shrink", type=int, default=0,
                    help="Orijinalin uzun kenarını bu piksele indir (0 = orijinali koru)")
    args = ap.parse_args()

    if not os.path.exists(args.file):
        sys.exit(f"❌  Dosya bulunamadı: {args.file}")

    print(f"📄  Kaynak: {os.path.basename(args.file)}")
    print("    Okunuyor… (gömülü görselli büyük dosyalarda birkaç dakika sürebilir)")
    wb = openpyxl.load_workbook(args.file, data_only=True)

    per_sheet = []
    total = 0
    for name in wb.sheetnames:
        imgs = sheet_images(wb[name])
        if imgs:
            per_sheet.append((name, imgs))
            total += len(imgs)
    print(f"🔎  {len(per_sheet)} sayfada {total} görsel bulundu.")
    for name, imgs in per_sheet:
        print(f"    {name}: {len(imgs)}")

    if args.dry_run:
        print("\n— ÖNİZLEME (hiçbir şey yazılmadı) —")
        print(f"    Kök klasör: {safe_name(args.root)}")
        print(f"    Orijinal: {'küçültülecek → ' + str(args.shrink) + 'px' if args.shrink else 'olduğu gibi korunacak'}")
        print(f"    Önizleme: {THUMB_MAX_PX}px, thumbs/ altında")
        return

    url, key, label = resolve_target(args.prod)
    print(f"🎯  Hedef: {label}")
    api = Api(url, key)

    ws_rows = api.get("/rest/v1/workspaces?select=id,name")
    if not ws_rows:
        sys.exit("❌  Çalışma alanı bulunamadı. Anahtar RLS'i aşmıyor olabilir "
                 "(Secret keys → default, sb_secret_…).")
    ws = next((w for w in ws_rows if w["name"] == "AF Operasyon"), ws_rows[0])
    workspace_id = ws["id"]
    print(f"✅  Çalışma alanı: {ws['name']} ({workspace_id})")

    # ── ÖN DENETİM: şema hazır mı? ───────────────────────────────────────────
    # Bu denetim OLMADAN betik görselleri önce depoya yüklüyor, sonra satırı
    # yazamayıp hata veriyordu: depoda sahipsiz dosyalar kalıyordu (Sıraç,
    # 2026-09-06 — prod'a migration uygulanmadan çalıştırıldı). Artık tek bir
    # bayt yüklenmeden önce durulur.
    required = ["thumb_path", "visibility", "section", "document_type", "owner_id"]
    missing = []
    for col in required:
        try:
            api.get(f"/rest/v1/operation_documents?select={col}&limit=1")
        except RuntimeError as e:
            if "PGRST204" in str(e) or "does not exist" in str(e) or "schema cache" in str(e):
                missing.append(col)
            else:
                raise
    if missing:
        sys.exit(
            "❌  Veritabanı şeması hazır değil — eksik sütun(lar): " + ", ".join(missing) + "\n"
            "\n    Bekleyen migration'lar bu hedefe UYGULANMAMIŞ. Hiçbir şey yüklenmedi.\n"
            "\n    Uygulamak için (projeye bağlıysanız):\n"
            "        supabase link --project-ref <proje-ref>\n"
            "        supabase db push\n"
            "\n    Ya da panelden: SQL Editor'e supabase/migrations/ altındaki\n"
            "    uygulanmamış dosyaların içeriğini sırayla yapıştırıp çalıştırın.\n"
            "\n    Sonra bu betiği tekrar çalıştırın — kaldığı yerden devam eder."
        )

    owner = api.get(f"/rest/v1/workspace_members?select=user_id&workspace_id=eq.{workspace_id}&role=eq.owner&limit=1")
    created_by = owner[0]["user_id"] if owner else None

    def ensure_folder(name, parent_id=None):
        """Klasörü bulur, yoksa açar. (workspace, parent, name) tekil."""
        q = (f"/rest/v1/document_folders?select=id&workspace_id=eq.{workspace_id}"
             f"&name=eq.{urllib.parse.quote(name)}")
        q += f"&parent_id=eq.{parent_id}" if parent_id else "&parent_id=is.null"
        found = api.get(q)
        if found:
            return found[0]["id"]
        body = {"workspace_id": workspace_id, "name": name, "parent_id": parent_id,
                "visibility": "all", "created_by": created_by}
        made = api.post("/rest/v1/document_folders", body, {"Prefer": "return=representation"})
        return made[0]["id"]

    # ── Eski yoldaki sahipsiz önizlemeleri temizle ──────────────────────────
    # İlk sürüm önizlemeyi "thumbs/<ws>/…" altına yazıyordu; depolama politikası
    # o yolu okuyamadığı için hepsi kırıktı ve doğru yola yeniden yazıldı. Eski
    # kopyalar kimsenin göremediği ama yer kaplayan dosyalar olarak kaldı.
    if args.drop_legacy_thumbs:
        print(f"🧹  Eski önizlemeler taranıyor: {LEGACY_THUMB_PREFIX}/{workspace_id}/")
        removed = 0
        while True:
            batch = api.list_objects(f"{LEGACY_THUMB_PREFIX}/{workspace_id}", limit=100)
            names = [o["name"] for o in (batch or []) if o.get("id")]
            if not names:
                break
            paths = [f"{LEGACY_THUMB_PREFIX}/{workspace_id}/{n}" for n in names]
            api.delete_objects(paths)
            removed += len(paths)
            print(f"    {removed} silindi…")
        print(f"✅  Eski önizleme temizliği bitti — {removed} dosya silindi.\n")
        if not args.file:
            return

    root_id = ensure_folder(safe_name(args.root))
    print(f"📁  Kök klasör hazır: {safe_name(args.root)}")

    uploaded = skipped = failed = repaired = 0
    bytes_original = bytes_thumb = 0

    for sheet_name, imgs in per_sheet:
        folder_id = ensure_folder(safe_name(sheet_name), root_id)
        for i, img in enumerate(imgs, 1):
            data = img["data"]
            ext = "jpg" if img["ext"] == "jpeg" else img["ext"]
            digest = hashlib.sha256(data).hexdigest()[:16]

            # İdempotenlik anahtarı DEPOLAMA YOLUDUR (içerik özetini o taşır),
            # dosya adı değil: dosya adı artık insan okusun diye sayfa adından
            # türetiliyor ve değişebilir.
            storage_path = f"{workspace_id}/{digest}.{ext}"
            human_name = f"{safe_name(sheet_name)} — görsel {i}.{ext}"
            existing = api.get(
                f"/rest/v1/operation_documents?select=id,file_name,thumb_path"
                f"&workspace_id=eq.{workspace_id}"
                f"&file_path=eq.{urllib.parse.quote(storage_path)}&limit=1"
            )
            if existing:
                row = existing[0]
                # ONARIM: eski koşulardan kalan kayıtlar özet adıyla ve yanlış
                # önizleme yoluyla duruyor olabilir. Yeniden yüklemeden düzeltilir.
                patch = {}
                if row.get("file_name") != human_name:
                    patch["file_name"] = human_name
                wants_thumb = f"{workspace_id}/{THUMB_DIR}/{digest}.{ext}"
                if row.get("thumb_path") != wants_thumb:
                    t = resize(data, ext, THUMB_MAX_PX)
                    if t:
                        try:
                            api.upload(wants_thumb, t, mimetypes.types_map.get("." + ext, f"image/{img['ext']}"))
                            patch["thumb_path"] = wants_thumb
                        except Exception as e:
                            print(f"    ⚠  önizleme onarılamadı ({sheet_name} #{i}): {e}")
                if patch:
                    try:
                        api._req("PATCH", f"/rest/v1/operation_documents?id=eq.{row['id']}",
                                 patch, {"Prefer": "return=minimal"})
                        repaired += 1
                    except Exception as e:
                        print(f"    ⚠  kayıt onarılamadı ({sheet_name} #{i}): {e}")
                else:
                    skipped += 1
                continue

            if args.shrink:
                smaller = resize(data, ext, args.shrink)
                if smaller and len(smaller) < len(data):
                    data = smaller

            thumb = resize(data, ext, THUMB_MAX_PX)
            path = storage_path
            thumb_path = f"{workspace_id}/{THUMB_DIR}/{digest}.{ext}" if thumb else None
            ctype = mimetypes.types_map.get("." + ext, f"image/{img['ext']}")

            try:
                api.upload(path, data, ctype)
                if thumb:
                    api.upload(thumb_path, thumb, ctype)
                api.post("/rest/v1/operation_documents", {
                    "workspace_id": workspace_id,
                    "title": f"{safe_name(sheet_name)} — görsel {i}",
                    # Sütun adları lib/actions/document-files.ts'teki yüklemeyle
                    # BİREBİR aynı: elle yüklenen dosyayla içe aktarılan dosya
                    # Drive'da ayırt edilemesin.
                    "document_type": "file",
                    "section": "teamwork",
                    "status": "approved",
                    "folder_id": folder_id,
                    "file_path": path,
                    "thumb_path": thumb_path,
                    # Ad İNSAN İÇİN: "K_15B_176 — görsel 12.jpg". Kimlik özeti
                    # dosya YOLUNDA duruyor, ekranda değil (Sıraç: "isimler
                    # neden bu şekilde?").
                    "file_name": human_name,
                    "file_size": len(data),
                    "file_mime": ctype,
                    "visibility": "all",
                    "owner_id": created_by,
                    "created_by": created_by,
                }, {"Prefer": "return=minimal"})
                uploaded += 1
                bytes_original += len(data)
                bytes_thumb += len(thumb) if thumb else 0
            except Exception as e:
                failed += 1
                print(f"    ⚠  {sheet_name} #{i}: {e}")

        print(f"    ✓ {sheet_name}  ({uploaded} yüklendi, {repaired} onarıldı, {skipped} atlandı)")

    mb = lambda n: f"{n / 1024 / 1024:.1f} MB"
    print(f"\n✅  Bitti — {uploaded} yüklendi · {repaired} onarıldı · {skipped} zaten vardı · {failed} başarısız")
    print(f"    Depolama: görseller {mb(bytes_original)} + önizlemeler {mb(bytes_thumb)}")
    if failed:
        print("    ⚠  Başarısız olanlar yukarıda listelendi; betiği tekrar çalıştırmak "
              "yalnız eksikleri dener (yüklenenler atlanır).")


if __name__ == "__main__":
    main()
