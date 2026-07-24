#!/usr/bin/env python3
"""Aslı Filinta üretim föyleri (Excel) → Supabase production_sheets.

Excel'deki her ürün föyünü (hücreler + gömülü görseller) uygulamaya aktarır.
Görseller Supabase Storage'a (production-sheets bucket) yüklenir; föy satırları
PostgREST ile eklenir. Aynı başlıklı mevcut föyler önce silinir (yeniden
çalıştırılabilir).

Env (yereller varsayılan):
  SUPABASE_URL        (vars: http://127.0.0.1:54321)
  SERVICE_ROLE_KEY    (zorunlu — yereli supabase status'tan)
  WORKSPACE_ID        (vars: 00000000-0000-0000-0000-000000000010)
  CREATED_BY          (vars: 00000000-0000-0000-0000-000000000002)
  UPDATED_BY          (vars: CREATED_BY)
  XLSX_PATH           (vars: uretim_foyu/Üretim Föyleri. 21 Temmuz 2026.xlsx)
  DRY_RUN=1           (yalnızca parse et, yaz — DB'ye dokunma)
"""
import os, sys, json, uuid, hashlib, re
from collections import Counter
import openpyxl
import urllib.request

URL = os.environ.get("SUPABASE_URL", "http://127.0.0.1:54321").rstrip("/")
KEY = os.environ.get("SERVICE_ROLE_KEY", "")
WS = os.environ.get("WORKSPACE_ID", "00000000-0000-0000-0000-000000000010")
CREATED_BY = os.environ.get("CREATED_BY", "00000000-0000-0000-0000-000000000002")
UPDATED_BY = os.environ.get("UPDATED_BY", CREATED_BY)
XLSX = os.environ.get("XLSX_PATH", "uretim_foyu/Üretim Föyleri. 21 Temmuz 2026.xlsx")
BUCKET = "production-sheets"
DRY = os.environ.get("DRY_RUN") == "1"

# Föy sayfaları (başlık = ÜRÜNÜN AÇIKLAMASI'ndan alınır). Föy olmayanlar hariç.
SKIP_SHEETS = {"G_RevizeFormu_1907", "Nakışçı ödeme listesi 0207"}


def s(v):
    if v is None:
        return ""
    t = str(v).strip()
    # openpyxl sayıları "1.0" olarak verir → tamsa "1" yap
    if re.fullmatch(r"-?\d+\.0", t):
        t = t[:-2]
    return t


def rows(ws):
    return [[s(c) for c in r] for r in ws.iter_rows(values_only=True)]


def after_colon(cell):
    return cell.split(":", 1)[1].strip() if ":" in cell else ""


def find_label(grid, *keys):
    """İlk hücresinde keys'ten biriyle başlayan satırı bul → (satır_idx, satır)."""
    for i, row in enumerate(grid):
        c0 = row[0] if row else ""
        for k in keys:
            if c0.upper().startswith(k.upper()):
                return i, row
    return -1, None


def parse_sheet(name, ws):
    grid = rows(ws)
    flat = {i: row for i, row in enumerate(grid)}

    def label_val(*keys):
        """A herhangi bir hücrede 'KEY:' ile başlayan → iki nokta sonrası."""
        for row in grid:
            for cell in row:
                for k in keys:
                    if cell.upper().startswith(k.upper()):
                        return after_colon(cell) or ""
        return ""

    kind = label_val("ÜRÜN CİNSİ")
    producer = label_val("ÜRETİCİ")
    delivery = label_val("TESLİM TARİHİ")
    production = label_val("ÜRETİM TARİHİ")
    season = label_val("SEZON")
    desc = label_val("ÜRÜNÜN AÇIKLAMASI")
    meterage = label_val("1 ÜRÜNE GİDEN METRAJ", "1 ÜRÜNE GIDEN METRAJ")
    title = desc or name

    # ── Ölçüler: "No" + "ÖLÇÜLER" başlığından sonra sayılı satırlar ──
    def collect_numbered(start_markers, stop_markers):
        out = []
        started = False
        for row in grid:
            joined = " ".join(row).upper()
            if not started:
                if any(m in joined for m in start_markers) and row and row[0].upper() == "NO":
                    started = True
                continue
            # dur: sonraki bölüm başlığı
            if any(m in joined for m in stop_markers):
                break
            no = row[0] if row else ""
            if not re.fullmatch(r"\d+", no):
                # boş numaralı satır ya da alakasız → numarasızsa dur değil, atla
                if no == "" and not any(row):
                    continue
                if not no.isdigit():
                    # bölüm bitmiş olabilir
                    if no and not no[0].isdigit():
                        break
                    continue
            vals = [c for c in row[1:] if c]
            label = vals[0] if len(vals) >= 1 else ""
            val = vals[1] if len(vals) >= 2 else ""
            if label:
                out.append({"no": no, "label": label, "val": val})
        return out

    meas_raw = collect_numbered(
        ["ÖLÇÜLER"], ["TESLİM EDİLEN", "TESLIM EDILEN"])
    measurements = [{"no": m["no"], "label": m["label"], "value": m["val"]} for m in meas_raw]

    deliv_raw = collect_numbered(
        ["TESLİM EDİLEN", "TESLIM EDILEN"],
        ["BEDEN DAĞILIMI", "BEDEN DAGILIMI", "ÜRETİM ADETİ", "URETIM ADETI", "YIKAMA"])
    delivered = [{"no": d["no"], "label": d["label"], "qty": d["val"]} for d in deliv_raw]

    # ── Beden dağılımı ──
    sizes, dist_rows = [], []
    for i, row in enumerate(grid):
        joined = " ".join(row).upper()
        if "BEDEN DAĞILIMI" in joined or "BEDEN DAGILIMI" in joined:
            # başlık satırındaki beden etiketleri (label hücresinden sonrakiler, TOPLAM hariç)
            header = row
            # ilk dolu olmayan/label hücresini atla
            hdr_vals = [c for c in header if c]
            # "BEDEN DAĞILIMI" veya "ÜRETİM ADETİ BEDEN DAĞILIMI" ilk eleman
            sizes = [c for c in hdr_vals[1:] if c.upper() not in ("TOPLAM",)]
            # sonraki satırları veri olarak al
            for r2 in grid[i + 1:]:
                jj = " ".join(r2).upper()
                if any(x in jj for x in ["YIKAMA", "KUMAŞ", "KUMAS", "SÜSLEME", "SUSLEME", "NOTLAR", "DİKİŞ", "DIKIS"]):
                    break
                vals = [c for c in r2 if c]
                if not vals:
                    continue
                # satır: [label?, v1, v2, ..., total?] — ilk hücre metinse label
                first = r2[0]
                if first and not re.fullmatch(r"[\d.,%=]+", first) and first.upper() != "":
                    label = first
                    values = [c for c in r2[1:1 + len(sizes)]]
                    total = r2[1 + len(sizes)] if len(r2) > 1 + len(sizes) else ""
                else:
                    label = "Üretim adeti"
                    values = [c for c in r2[:len(sizes)] if True][:len(sizes)]
                    total = r2[len(sizes)] if len(r2) > len(sizes) else ""
                total = "" if total.startswith("=") else total
                if any(values) or total:
                    dist_rows.append({"label": label, "values": values, "total": total})
            break
    size_distribution = {"sizes": sizes, "rows": dist_rows} if sizes else {"sizes": [], "rows": []}

    # ── Uzun metin bölümleri (label satırında 2. hücreden itibaren) ──
    def section_text(*keys):
        parts = []
        for row in grid:
            for j, cell in enumerate(row):
                if any(cell.upper().startswith(k.upper()) for k in keys):
                    rest = [c for c in row[j + 1:] if c]
                    if rest:
                        parts.append(" ".join(rest))
                    else:
                        # değer iki nokta sonrasında olabilir
                        av = after_colon(cell)
                        if av:
                            parts.append(av)
        return "\n".join(parts).strip() or None

    wash = section_text("YIKAMA TALİMATI", "YIKAMA TALIMATI")
    fabric_lining = section_text("KUMAŞ / ASTAR", "KUMAS / ASTAR")
    fabric_info = section_text("KUMAŞ BİLGİSİ", "KUMAS BILGISI", "KUMAŞ AÇIKLAMASI", "KUMAS ACIKLAMASI")
    accessories = section_text("AKSESUARLAR BİLGİSİ", "AKSESUARLAR BILGISI")
    embell = section_text("SÜSLEMELER VE AKSESUAR AÇIKLAMASI", "SUSLEMELER VE AKSESUAR ACIKLAMASI")
    sewing = section_text("DİKİŞ TALİMATI", "DIKIS TALIMATI", "NOTLAR")
    workmanship = section_text("ÖZEL İŞÇİLİK NOTLARI", "OZEL ISCILIK NOTLARI")
    qc = section_text("KALİTE KONTROL REVIZYON TARIHI", "KALITE KONTROL REVIZYON")
    revision = section_text("REVIZYON NOTLARI")
    waste = section_text("ÜRETİM FİRE PAYI", "URETIM FIRE PAYI")

    return {
        "title": title, "status": "active", "product_kind": kind or None,
        "producer": producer or None, "description": desc or None,
        "season": season or None, "production_date": production or None,
        "delivery_date": delivery or None, "meterage": meterage or None,
        "measurements": measurements, "delivered_items": delivered,
        "size_distribution": size_distribution,
        "wash_instruction": wash, "fabric_lining": fabric_lining,
        "fabric_info": fabric_info, "accessories_info": accessories,
        "embellishments": embell, "sewing_instruction": sewing,
        "workmanship_notes": workmanship, "qc_revision": qc,
        "revision_notes": revision, "production_waste": waste,
    }


def read_sheet_images(ws):
    """Sayfadaki görselleri BİR KEZ okur (openpyxl _data() ikinci okumada kapanır)."""
    out = []
    for im in getattr(ws, "_images", []):
        try:
            data = im._data()
        except Exception:
            continue
        try:
            anc = im.anchor._from
            row1 = anc.row + 1
        except Exception:
            row1 = 99
        fmt = (getattr(im, "format", None) or "jpeg").lower()
        if fmt == "jpg":
            fmt = "jpeg"
        out.append({
            "data": data, "hash": hashlib.md5(data).hexdigest()[:8],
            "ext": fmt, "row": row1,
        })
    return out


def is_foy(grid):
    """İlk 6 satırda föy işareti var mı?"""
    for row in grid[:6]:
        joined = " ".join(row).upper()
        if "ÜRETİM FÖYÜ" in joined or "ÜRÜN CİNSİ" in joined or "ÜRÜNÜN AÇIKLAMASI" in joined:
            return True
    return False


def section_for_text(t):
    """Bir satır metnini editör bölüm enum'una eşler (öncelik sırası önemli)."""
    t = t.upper()
    if "TEKNİK ÇİZİM" in t or "ÖLÇÜLER" in t:
        return "technical_drawing"
    if "SÜSLEME" in t:                      # SÜSLEMELER (VE AKSESUAR) → süsleme
        return "embellishments"
    if "AKSESUAR" in t:                     # AKSESUARLAR BİLGİSİ → aksesuar
        return "accessories"
    if "DİKİŞ TALİMATI" in t or "NOTLAR" in t:
        return "sewing"
    if "KUMAŞ" in t or "ASTAR" in t or "YIKAMA" in t:  # kumaş/astar + beden foto
        return "fabric"
    return None


def build_section_markers(grid):
    """(1-tabanlı satır, bölüm) işaretlerini sıralı döndürür."""
    markers = []
    for i, row in enumerate(grid):
        joined = " ".join(c for c in row if c)
        sec = section_for_text(joined)
        if sec:
            markers.append((i + 1, sec))
    markers.sort(key=lambda m: m[0])
    return markers


def assign_section(markers, img_row):
    """Görselin bulunduğu satırın ÜSTÜNDEKİ en yakın bölüm başlığına ata."""
    chosen = "technical_drawing"  # ilk başlığın da üstündeyse teknik çizim alanı
    for row, sec in markers:
        if row <= img_row:
            chosen = sec
        else:
            break
    return chosen


def http(method, path, body=None, headers=None, raw=False):
    url = URL + path
    h = {"apikey": KEY, "Authorization": f"Bearer {KEY}"}
    if headers:
        h.update(headers)
    data = body if raw else (json.dumps(body).encode() if body is not None else None)
    if not raw and body is not None:
        h["Content-Type"] = "application/json"
    req = urllib.request.Request(url, data=data, headers=h, method=method)
    with urllib.request.urlopen(req) as resp:
        return resp.status, resp.read()


def main():
    wb = openpyxl.load_workbook(XLSX)
    # Görselleri sayfa bazında TEK KEZ oku (openpyxl _data() ikinci okumada kapanır).
    # NOT: Artık hash-dedup YOK — tekrar eden gerçek fotoğraflar (ör. AF marka
    # etiketi birden fazla föyde) korunur. Yalnızca en üstteki (satır ≤ 2) logo
    # atlanır; bölüm ataması Excel'deki en yakın başlığa göre yapılır.
    sheet_imgs = {name: read_sheet_images(wb[name]) for name in wb.sheetnames}

    foys = []
    seen_titles = {}
    for name in wb.sheetnames:
        if name in SKIP_SHEETS:
            continue
        ws = wb[name]
        grid = rows(ws)
        if not is_foy(grid):
            print(f"  atlandı (föy değil): {name}")
            continue
        parsed = parse_sheet(name, ws)
        # Başlık çakışması (kaynakta kopyala-yapıştır hatası) → sayfa adıyla ayır.
        title = parsed["title"]
        clean_sheet = re.sub(r"\s*Föyü$", "", name).strip()
        if title in seen_titles or not title:
            title = clean_sheet or f"{title} ({name})"
        seen_titles[title] = True
        parsed["title"] = title
        # Görselleri: en üstteki logoyu (satır ≤ 2) atla; bölümü en yakın
        # başlığa göre ata (Excel'deki gibi ilgili başlığın altına düşer).
        markers = build_section_markers(grid)
        imgs = []
        for im in sheet_imgs.get(name, []):
            if im["row"] <= 2:  # başlık şeridindeki AF logosu
                continue
            imgs.append({
                "data": im["data"], "ext": im["ext"],
                "section": assign_section(markers, im["row"]),
            })
        parsed["_images"] = imgs
        foys.append(parsed)
        print(f"  ✓ {name} → '{parsed['title']}' | ölçü:{len(parsed['measurements'])} "
              f"teslim:{len(parsed['delivered_items'])} beden:{len(parsed['size_distribution']['rows'])} "
              f"görsel:{len(imgs)}")

    if DRY:
        # ilk föyü detaylı yaz
        sample = {k: v for k, v in foys[0].items() if k != "_images"}
        print("\n=== ÖRNEK PARSE (ilk föy) ===")
        print(json.dumps(sample, ensure_ascii=False, indent=2))
        print(f"\nToplam {len(foys)} föy parse edildi (DRY_RUN — DB'ye yazılmadı).")
        return

    if not KEY:
        print("HATA: SERVICE_ROLE_KEY gerekli.", file=sys.stderr)
        sys.exit(1)

    # WORKSPACE_ID / CREATED_BY verilmediyse otomatik tespit (canlıda kolaylık —
    # Supabase bulutunda service_role tam erişime sahiptir). Başarısız olursa
    # açık env vermeye yönlendirir.
    global WS, CREATED_BY, UPDATED_BY
    if not os.environ.get("WORKSPACE_ID"):
        try:
            _, body = http("GET", "/rest/v1/workspaces?select=id,name&limit=2")
            wss = json.loads(body)
            assert wss
            WS = wss[0]["id"]
            note = f" ({wss[0].get('name','')})" + (
                f"  [DİKKAT: birden fazla workspace var, ilki seçildi — doğruysa devam, değilse WORKSPACE_ID verin]"
                if len(wss) > 1 else "")
            print(f"Otomatik workspace: {WS}{note}")
        except Exception as e:
            print(f"HATA: workspace otomatik bulunamadı ({e}). WORKSPACE_ID env verin.", file=sys.stderr)
            sys.exit(1)
    if not os.environ.get("CREATED_BY"):
        try:
            _, body = http("GET", f"/rest/v1/workspace_members?workspace_id=eq.{WS}&order=role.asc&select=user_id,role&limit=1")
            mem = json.loads(body)
            assert mem
            CREATED_BY = mem[0]["user_id"]
            UPDATED_BY = CREATED_BY
            print(f"Otomatik created_by: {CREATED_BY} (rol: {mem[0].get('role')})")
        except Exception as e:
            print(f"HATA: workspace üyesi otomatik bulunamadı ({e}). CREATED_BY env verin.", file=sys.stderr)
            sys.exit(1)

    for f in foys:
        images = f.pop("_images")
        title = f["title"]
        # aynı başlıklı mevcut föyleri sil (idempotent)
        q = f"/rest/v1/production_sheets?workspace_id=eq.{WS}&title=eq.{urllib.parse.quote(title)}"
        try:
            http("DELETE", q, headers={"Prefer": "return=minimal"})
        except Exception as e:
            print(f"  uyarı silme: {e}")

        sheet_id = str(uuid.uuid4())
        photo_refs = []
        for img in images:
            path = f"{WS}/{sheet_id}/{uuid.uuid4()}.{img['ext']}"
            ctype = f"image/{img['ext']}"
            try:
                http("POST", f"/storage/v1/object/{BUCKET}/{path}", body=img["data"],
                     headers={"Content-Type": ctype, "x-upsert": "true"}, raw=True)
                pub = f"{URL}/storage/v1/object/public/{BUCKET}/{path}"
                photo_refs.append({"url": pub, "path": path, "section": img["section"]})
            except Exception as e:
                print(f"  uyarı görsel yükleme ({title}): {e}")

        row = {
            "id": sheet_id, "workspace_id": WS, "created_by": CREATED_BY,
            "updated_by": UPDATED_BY, "photo_refs": photo_refs, **f,
        }
        try:
            st, _ = http("POST", "/rest/v1/production_sheets", body=row,
                         headers={"Prefer": "return=minimal"})
            print(f"  ↑ eklendi: {title} (görsel {len(photo_refs)}) [{st}]")
        except Exception as e:
            print(f"  HATA ekleme ({title}): {e}")

    print(f"\nBitti: {len(foys)} föy aktarıldı → {URL}")


if __name__ == "__main__":
    import urllib.parse
    main()
